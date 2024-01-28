
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_workbook():
    workbook = Workbook()
    return workbook

def create_worksheet(workbook, sheet_title="Sheet1"):
    worksheet = workbook.create_sheet(title=sheet_title)
    return worksheet


def read_or_create_workbook(file_path):
    # Create a new workbook if the file doesn't exist
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    return wb

def read_or_create_worksheet(wb, sheet_name):
    # Check if the sheet exists, create it if not
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]

    return ws

def next_sheet_row(ws):
    # Get the maximum row number in the existing sheet
    max_row = ws.max_row

     # Calculate the next row
    next_row = 1 if max_row == 1 else max_row + 2

    return next_row
