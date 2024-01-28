import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


from excel_builder import read_or_create_workbook
from excel_builder import read_or_create_worksheet
from excel_builder import next_sheet_row


def first_row_to_string(data):
    # Convert the top row to strings
    data[0] = [str(item) for item in data[0]]

    return data

def get_excel_cell(col, row):
    # Convert column number to Excel column letter
    col_letter = chr(ord('A') + col - 1)

    # Create Excel cell reference
    cell_reference = f"{col_letter}{row}"

    return cell_reference


def excel_table(
        ws,
        start_col,
        start_row,
        end_col,
        end_row,
        displayName="Table1",
):
    start_cell = get_excel_cell(start_col,start_row,)
    end_cell = get_excel_cell(end_col,end_row,)

    tab = Table(displayName=displayName, ref=f"{start_cell}:{end_cell}")
    
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    tab.tableStyleInfo = style
    
    '''
    Table must be added using ws.add_table() method to avoid duplicate names.
    Using this method ensures table name is unque through out defined names and all other table name. 
    '''

    ws.add_table(tab)

def add_table(
    df: pd.DataFrame,
    sheet_name,
    displayName,
    file_path='tst.xlsx',    
):
    # Convert column names to strings
    df.columns = df.columns.astype(str)

    # Create a new workbook if the file doesn't exist
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    # Check if the sheet exists, create it if not
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]


    # Get the maximum row number in the existing sheet
    next_row = ws.next_row
    if next_row != 1: next_row=next_row+2

    # Write the headers to the Excel sheet
    for col_num, value in enumerate(df.columns, 1):
        ws.cell(row=(next_row), column=col_num, value=value)

    # Write the data to the Excel sheet
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=(next_row+row_num-1), column=col_num, value=value)

    excel_table(
            ws=ws,
            start_col=1,
            start_row=next_row,
            end_col=df.shape[1],
            end_row=(next_row+df.shape[0]),
            displayName=displayName,
    )

    # Save the changes to the existing workbook
    wb.save(file_path)


def add_df_data(
    df: pd.DataFrame,
    sheet_name,
    file_path='tst.xlsx',    
):
    # Convert column names to strings
    df.columns = df.columns.astype(str)

    # Create a new workbook if the file doesn't exist
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    # Check if the sheet exists, create it if not
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb[sheet_name]


    # Get the maximum row number in the existing sheet
    next_row = ws.next_row
    if next_row != 1: next_row=next_row+2

    # Write the headers to the Excel sheet
    for col_num, value in enumerate(df.columns, 1):
        ws.cell(row=(next_row), column=col_num, value=value)

    # Write the data to the Excel sheet
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=(next_row+row_num-1), column=col_num, value=value)

    # Save the changes to the existing workbook
    wb.save(file_path)

def add_df_data_as_table(
    df: pd.DataFrame,
    sheet_name,
    displayName,
    file_path='tst.xlsx',    
):
    # Convert column names to strings
    df.columns = df.columns.astype(str)

    # Create a new workbook if the file doesn't exist
    wb = read_or_create_workbook(file_path=file_path)

    # Creaate sheet
    ws = read_or_create_worksheet(wb=wb, sheet_name=sheet_name)

    # Next row to use
    next_row=next_sheet_row(ws=ws)

    # Write the headers to the Excel sheet
    for col_num, value in enumerate(df.columns, 1):
        ws.cell(row=(next_row), column=col_num, value=value)

    # Write the data to the Excel sheet
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=(next_row+row_num-1), column=col_num, value=value)

    excel_table(
            ws=ws,
            start_col=1,
            start_row=next_row,
            end_col=df.shape[1],
            end_row=(next_row+df.shape[0]),
            displayName=displayName,
    )

    # Save the changes to the existing workbook
    wb.save(file_path)